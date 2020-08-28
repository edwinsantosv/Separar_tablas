import pandas as pd
import os
import re
#Si tienes dudas sobre este package, revisar https://realpython.com/openpyxl-excel-spreadsheets-python/
from openpyxl import load_workbook
from openpyxl import Workbook

#1. CREACIÓN DE LA TABLA RESUMEN

try:
    TablaRes=Workbook()
    HojaLN=TablaRes.active
    TablaRes.save(filename="TablaResumen.xlsx")
except:
    pass

#2. LECTURA DE ARCHIVOS

#filepath en donde están los archivos
fpathtablas='tablas_excel'

#Esta función se encarga de leer y generar una lista con las tablas
Lista_tablas=os.listdir(fpathtablas)

#Se declara un contador para la lectura de los archivos
i=0

#Se declara un contador para la escritura en la hoja
i1=0

#Número de archivos
numarch=len(Lista_tablas)

#Para los siguientes pasos se pasará a utilizar un loop
while i<numarch:

#Nombre del archivo ejem:"tabla 1.xlsx"
    Nomarch=Lista_tablas[i]

#Nombre del archivo sin .xlsx ejem:"tabla 1"
    Nomsolo=re.findall('([^_]*)\.',Nomarch)[0] #para la expresión regular ([^_]*)\., quiere decir,
                                               #todo hasta que te encuentres con un punto
    i=i+1

#3. LECTURA DE CELDAS

#Recuerda que la tabla está dentro del folder tablas_excel\
    workbook=load_workbook(filename='tablas_excel\\'+Nomarch)
    Hoja1=workbook.active

#Se ubica cada uno de los datos de las tablas
    NomAp=Hoja1["A5"].value
    DNI=Hoja1["A7"].value
    Dir=Hoja1["A9"].value
    Dist=Hoja1["A11"].value
    Tel=Hoja1["A13"].value
    Ant=Hoja1["A15"].value
    Sint=Hoja1["A31"].value

    DProd=Hoja1["E5"].value
    Prod=Hoja1["E7"].value
    Numser=Hoja1["E9"].value
    FechaCom=Hoja1["E11"].value

#4. ESCRITURA DE CELDAS EN LA HOJA RESUMEN
#y se escriben en la hoja nueva
    try:
        HojaLN["A"+str(i1+1)]=NomAp
        HojaLN["B"+str(i1+1)]=DNI
        HojaLN["C"+str(i1+1)]=Dir
        HojaLN["D"+str(i1+1)]=Dist
        HojaLN["E"+str(i1+1)]=Tel
        HojaLN["F"+str(i1+1)]=Ant
        HojaLN["G"+str(i1+1)]=Sint
        HojaLN["H" + str(i1+1)] =DProd
        HojaLN["I" + str(i1+1)] =Prod
        HojaLN["J" + str(i1+1)] = Numser
        HojaLN["K" + str(i1+1)] = FechaCom
    except: pass
    TablaRes.save(filename="TablaResumen.xlsx")

    i1=i1+1